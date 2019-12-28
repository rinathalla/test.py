import openpyxl
wb = openpyxl.Workbook()
wb
wb.get_sheet_names()
sheet=wb.get_sheet_by_name('Sheet')
sheet['A1'].value
sheet['A1']='Rina'#to add value to column

import os
#to save to harddrive
os.chdir('C:\\Users\\Administrator\\PycharmProjects\\Project1\\Screenshots')
wb.save('example.xlsx') #Created copy of original file
sheet2= wb.create_sheet() # creating a new sheet
sheet2.title="Newsheet"
wb.get_sheet_names()
['Sheet', 'Sheet1', 'Newsheet']
wb.save('example.xlsx')
wb.create_sheet(index=0,title='1SheetRenamed') #Renaming 1st sheet
#<Worksheet "1SheetRenamed">
wb.save('example.xlsx')
